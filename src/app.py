#!/usr/bin/env python3
"""
Enhanced Flask backend for MIL-HDBK-217F Reliability Prediction Calculator
Project-based application with improved functionality
"""

import os
import sqlite3
import math
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from flask import send_file
from datetime import datetime, timezone, timedelta
from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from config import Config

# Initialize Flask app
app = Flask(__name__)
app.config.from_object(Config)

def init_database():
    """Initialize database if it doesn't exist"""
    database_path = Config.get_database_path()
    
    if not os.path.exists(database_path):
        print(f"Database not found at {database_path}. Creating new database...")
        os.makedirs(os.path.dirname(database_path), exist_ok=True)
        
        try:
            # Import and run database initialization
            import sys
            database_dir = os.path.join(Config.BASE_DIR, 'database')
            if database_dir not in sys.path:
                sys.path.append(database_dir)
            
            from database.init_db import create_database
            create_database()
            print(f"Database created successfully at: {database_path}")
        except ImportError as e:
            print(f"Error importing init_db: {e}")
            # Create database manually if import fails
            create_database_manually(database_path)

def create_database_manually(db_path):
    """Fallback database creation"""
    try:
        import sqlite3
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Basic capacitor_styles table
        cursor.execute('''
        CREATE TABLE capacitor_styles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            style TEXT NOT NULL,
            spec_number TEXT,
            description TEXT,
            lambda_b REAL NOT NULL,
            pi_t_column INTEGER,
            pi_c_column INTEGER,
            pi_v_column INTEGER,
            pi_sr REAL DEFAULT 1.0,
            created_at TIMESTAMP DEFAULT (datetime('now', '+7 hours'))
            )
            ''')
    
        # Temperature factor table
        cursor.execute('''
        CREATE TABLE temperature_factors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            temperature INTEGER NOT NULL,
            column_1 REAL,
            column_2 REAL,
            created_at TIMESTAMP DEFAULT (datetime('now', '+7 hours'))
        )
        ''')
        
        # Capacitance factor table
        cursor.execute('''
        CREATE TABLE capacitance_factors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            capacitance REAL NOT NULL,
            column_1 REAL,
            column_2 REAL,
            created_at TIMESTAMP DEFAULT (datetime('now', '+7 hours'))
        )
        ''')
        
        # Voltage stress factor table
        cursor.execute('''
        CREATE TABLE voltage_stress_factors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            voltage_stress REAL NOT NULL,
            column_1 REAL,
            column_2 REAL,
            column_3 REAL,
            column_4 REAL,
            column_5 REAL,
            created_at TIMESTAMP DEFAULT (datetime('now', '+7 hours'))
        )
        ''')
        
        # Quality factor table
        cursor.execute('''
        CREATE TABLE quality_factors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            quality_level TEXT NOT NULL,
            pi_q REAL NOT NULL,
            created_at TIMESTAMP DEFAULT (datetime('now', '+7 hours'))
        )
        ''')
        
        # Environment factor table
        cursor.execute('''
        CREATE TABLE environment_factors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            environment TEXT NOT NULL,
            pi_e REAL NOT NULL,
            created_at TIMESTAMP DEFAULT (datetime('now', '+7 hours'))
        )
        ''')
        
        # Series resistance factor table (for tantalum capacitors)
        cursor.execute('''
        CREATE TABLE series_resistance_factors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            resistance_range TEXT NOT NULL,
            pi_sr REAL NOT NULL,
            created_at TIMESTAMP DEFAULT (datetime('now', '+7 hours'))
        )
        ''')
        
        # Insert basic data
        cursor.execute('''
            INSERT OR REPLACE INTO capacitor_styles 
            (style, spec_number, description, lambda_b, pi_t_column, pi_c_column, pi_v_column, pi_sr)
            VALUES ('CP', '25', 'Capacitor, Fixed, Paper-Dielectric, Direct Current (Hermetically Sealed in Metal Cases)', 0.00037, 1, 1, 1, 1.0),
            ('CA', '12889', 'Capacitor, By-Pass, Radio Interference Reduction, Paper Dielectric, AC and DC (Hermetically sealed in Metallic Case)', 0.00037, 1, 1, 1, 1.0),
            ('CZ, CZR', '11693', 'Capacitor, Feed through, Radio Interference Reduction AC and DC (Hermetically sealed in metal cases), Established and Nonestablished Reliability', 0.00037, 1, 1, 1, 1.0),
            ('CQ, CQR', '19978', 'Capacitor, Fixed Plastic (or Paper-Plastic) Dielectric (Hermetically sealed in metal, ceramic, or glass cases), Established and Nonestablished Reliability', 0.00051, 1, 1, 1, 1.0),
            ('CH', '18312', 'Capacitor, Fixed, Metallized (Paper, Paper Plastic or Plastic Film) Dielectric, Direct Current (Hermetically Sealed in Metal Cases)', 0.00037, 1, 1, 1, 1.0),
            ('CHR', '39022', 'Capacitor, Fixed, Metallized Paper, Paper-Plastic Film or Plastic Film Dielectric', 0.00051, 1, 1, 1, 1.0),
            ('CFR', '55514', 'Capacitor, Fixed, Plastic (or Metallized Plastic) Dielectric, Direct Current in Non-Metal Cases', 0.00051, 1, 1, 1, 1.0),
            ('CRH', '83421', 'Capacitor, Fixed Supermetallized Plastic Film Dielectric (DC, AC or DC and AC) Hermetically sealed Established Reliability', 0.00051, 1, 1, 1, 1.0),
            ('CM', '5', 'Capacitor, Fixed, Mica Dielectric', 0.00076, 2, 1, 2, 1.0),
            ('CMR', '39001', 'Capacitor, Fixed, Mica Dielectric, Established Reliability', 0.00076, 2, 1, 2, 1.0),
            ('CB', '10950', 'Capacitor, Fixed, Mica Dielectric, Button Style', 0.00076, 2, 1, 2, 1.0),
            ('CY', '11272', 'Capacitor, Fixed, Glass Dielectric', 0.00076, 2, 1, 2, 1.0),
            ('CYR', '23269', 'Capacitor, Fixed, Glass Dielectric, Established Reliability', 0.00076, 2, 1, 2, 1.0),
            ('CK', '11015', 'Capacitor, Fixed, Ceramic Dielectric (General Purpose)', 0.00099, 2, 1, 3, 1.0),
            ('CKR', '39014', 'Capacitor, Fixed, Ceramic Dielectric (General Purpose), Established Reliability', 0.00099, 2, 1, 3, 1.0),
            ('CC, CCR', '20', 'Capacitor, Fixed, Ceramic Dielectric (Temperature Compensating), Established and Nonestablished Reliability', 0.00099, 2, 1, 3, 1.0),
            ('CDR', '55681', 'Capacitor, Chip, Multiple Layer, Fixed, Ceramic Dielectric, Established Reliability', 0.0020, 2, 1, 3, 1.0),
            ('CSR', '39003', 'Capacitor, Fixed, Electrolytic (Solid Electrolyte), Tantalum, Established Reliability', 0.00040, 1, 2, 4, 1.0),
            ('CWR', '55365', 'Capacitor, Fixed, Electrolytic (Tantalum), Chip, Established Reliability', 0.00005, 1, 2, 4, 1.0),
            ('CL', '3965', 'Capacitor, Fixed, Electrolytic (Nonsolid Electrolyte), Tantalum', 0.00040, 1, 2, 4, 1.0),
            ('CLR', '39006', 'Capacitor, Fixed, Electrolytic (Nonsolid Electrolyte), Tantalum, Established Reliability', 0.00040, 1, 2, 4, 1.0),
            ('CRL', '83500', 'Capacitor, Fixed, Electrolytic (Nonsolid Electrolyte), Tantalum Cathode', 0.00040, 1, 2, 4, 1.0),
            ('CU, CUR', '39018', 'Capacitor, Fixed, Electrolytic (Aluminum Oxide), Established Reliability and Nonestablished Reliability', 0.00012, 2, 2, 1, 1.0),
            ('CE', '62', 'Capacitor, Fixed Electrolytic (DC, Aluminum, Dry Electrolyte, Polarized)', 0.00012, 2, 2, 1, 1.0),
            ('CV', '81', 'Capacitor, Variable, Ceramic Dielectric (Trimmer)', 0.0079, 1, 1, 5, 1.0),
            ('PC', '14409', 'Capacitor, Variable (Piston Type, Tubular Trimmer)', 0.0060, 2, 1, 5, 1.0),
            ('CT', '92', 'Capacitor, Variable, Air Dielectric (Trimmer)', 0.0000072, 2, 1, 5, 1.0),
            ('CG', '23183', 'Capacitor, Fixed or Variable, Vacuum Dielectric', 0.0060, 1, 1, 5, 1.0)
        ''')
        
        cursor.execute(''' 
            INSERT OR REPLACE INTO temperature_factors
            (temperature, column_1, column_2)
            VALUES  (20, 0.91, 0.79),
                    (30, 1.1, 1.3),
                    (40, 1.3, 1.9),
                    (50, 1.6, 2.9),
                    (60, 1.8, 4.2),
                    (70, 2.2, 6.0),
                    (80, 2.5, 8.4),
                    (90, 2.8, 11),
                    (100, 3.2, 15),
                    (110, 3.7, 21),
                    (120, 4.1, 27),
                    (130, 4.6, 35),
                    (140, 5.1, 44),
                    (150, 5.6, 56)
        ''')

        cursor.execute(''' 
            INSERT OR REPLACE INTO capacitance_factors 
            (capacitance, column_1, column_2)
            VALUES  (0.000001, 0.29, 0.04),
                    (0.00001, 0.35, 0.07),
                    (0.0001, 0.44, 0.12),
                    (0.001, 0.54, 0.20),
                    (0.01, 0.66, 0.35),
                    (0.05, 0.76, 0.50),
                    (0.1, 0.81, 0.59),
                    (0.5, 0.94, 0.85),
                    (1.0, 1.0, 1.0),
                    (3.0, 1.1, 1.3),
                    (10.0, 1.2, 1.6),
                    (30.0, 1.3, 1.9),
                    (100.0, 1.4, 2.3),
                    (300.0, 1.6, 3.4),
                    (1000.0, 1.9, 4.9),
                    (3000.0, 2.1, 6.3),
                    (10000.0, 2.3, 8.3),
                    (30000.0, 2.5, 11),
                    (60000.0, 2.7, 13),
                    (120000.0, 2.9, 15)           
        ''')

        cursor.execute(''' 
            INSERT OR REPLACE INTO voltage_stress_factors 
        (voltage_stress, column_1, column_2, column_3, column_4, column_5)
        VALUES  (0.1, 1.0, 1.0, 1.0, 1.0, 1.0),
                (0.2, 1.0, 1.0, 1.0, 1.0, 1.1),
                (0.3, 1.0, 1.0, 1.1, 1.0, 1.2),
                (0.4, 1.1, 1.0, 1.3, 1.0, 1.5),
                (0.5, 1.4, 1.2, 1.6, 1.0, 2.0),
                (0.6, 2.0, 2.0, 2.0, 2.0, 2.7),
                (0.7, 3.2, 5.7, 2.6, 15, 3.7),
                (0.8, 5.2, 19, 3.4, 130, 5.1),
                (0.9, 8.6, 59, 4.4, 990, 6.8),
                (1.0, 14, 166, 5.6, 5900, 9.0)
        ''')

        cursor.execute(''' 
            INSERT OR REPLACE INTO quality_factors (quality_level, pi_q)
        VALUES  ('D', 0.001),
                ('C', 0.01),
                ('S,B', 0.03),
                ('R', 0.1),
                ('P', 0.3),
                ('M', 1.0),
                ('L', 1.5),
                ('Non-Established Reliability', 3.0),
                ('Commercial or Unknown', 10.0)
        ''')

        cursor.execute(''' 
            INSERT OR REPLACE INTO environment_factors (environment, pi_e)
        VALUES  ('GB', 1.0),
                ('GF', 10),
                ('GM', 20),
                ('NS', 7.0),
                ('NU', 15),
                ('AIC', 12),
                ('AIF', 15),
                ('AUC', 25),
                ('AUF', 30),
                ('ARW', 40),
                ('SF', 0.50),
                ('MF', 20),
                ('ML', 50),
                ('CL', 570)
        ''')

        cursor.execute(''' 
            INSERT OR REPLACE INTO series_resistance_factors (resistance_range, pi_sr)
        VALUES  ('>0.8', 0.66),
                ('>0.6 to 0.8', 1.0),
                ('>0.4 to 0.6', 1.3),
                ('>0.2 to 0.4', 2.0),
                ('>0.1 to 0.2', 2.7),
                ('0 to 0.1', 3.3)
        ''')

        conn.commit()
        conn.close()
        print(f"Basic database created at: {db_path}")
    except Exception as e:
        print(f"Error creating database manually: {e}")

def get_db_connection():
    """Get database connection with proper path resolution"""
    database_path = Config.get_database_path()
    
    if not os.path.exists(database_path):
        init_database()
    
    conn = sqlite3.connect(database_path)
    conn.row_factory = sqlite3.Row
    return conn

def calculate_temperature_factor(temperature, column):
    """Calculate temperature factor using MIL-HDBK-217F equation"""
    if column == 1:
        ea = Config.TEMP_FACTOR_EA_COLUMN1
    else:
        ea = Config.TEMP_FACTOR_EA_COLUMN2
    
    # Convert Celsius to Kelvin
    temp_kelvin = temperature + 273
    
    # π_T = exp(-Ea/k * (1/T - 1/T_ref))
    factor = math.exp(-ea / Config.BOLTZMANN_CONSTANT * 
                     (1/temp_kelvin - 1/Config.REFERENCE_TEMP))
    return round(factor, 6)

def calculate_capacitance_factor(capacitance, column):
    """Calculate capacitance factor using MIL-HDBK-217F equation"""
    if capacitance <= 0:
        return 1.0
    
    if column == 1:
        exponent = Config.CAP_FACTOR_EXP_COLUMN1
    else:
        exponent = Config.CAP_FACTOR_EXP_COLUMN2
    
    # π_C = C^exponent
    factor = capacitance ** exponent
    return round(factor, 6)

def get_exact_or_calculate_factor(value, data_points, value_column, factor_column, calculation_type, column_number=None):
    """Get exact value from table or calculate using equation"""
    
    # First, check if exact value exists in table
    for row in data_points:
        if abs(row[value_column] - value) < 1e-10:
            return row[factor_column]
    
    # If not found in table, use appropriate calculation
    if calculation_type == 'temperature':
        return calculate_temperature_factor(value, column_number)
    elif calculation_type == 'capacitance':
        return calculate_capacitance_factor(value, column_number)
    elif calculation_type == 'voltage_stress':
        if column_number == 1:
            return (value / 0.6) ** 5 + 1 
        elif column_number == 2:
            return (value / 0.6) ** 10 + 1 
        elif column_number == 3:
            return (value / 0.6) ** 3 + 1 
        elif column_number == 4:
            return (value / 0.6) ** 17 + 1 
        elif column_number == 5:
            return (value / 0.5) ** 3 + 1 
        else:
            return 1.0
    elif calculation_type == 'resistor_temperature':
        return calculate_resistor_temperature_factor(value, column_number)
    elif calculation_type == 'resistor_power':
        return calculate_resistor_power_factor(value)
    elif calculation_type == 'resistor_stress':
        return calculate_resistor_stress_factor(value, column_number)
    elif calculation_type == 'inductor_temperature':
        return calculate_inductor_temperature_factor(value)
    else:
        return 1.0
# Routes
@app.route('/')
def index():
    """Main application page"""
    return render_template('index.html')

@app.route('/splash')
def splash():
    """Splash screen for desktop application"""
    return render_template('splash.html')

@app.route('/api/capacitor-styles')
def get_capacitor_styles():
    """Get all capacitor styles"""
    try:
        conn = get_db_connection()
        styles = conn.execute('''
            SELECT style, spec_number, description, lambda_b, 
                   pi_t_column, pi_c_column, pi_v_column, pi_sr
            FROM capacitor_styles 
            ORDER BY style
        ''').fetchall()
        conn.close()
        
        return jsonify([dict(row) for row in styles])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/quality-levels')
def get_quality_levels():
    """Get all quality levels"""
    try:
        conn = get_db_connection()
        qualities = conn.execute('''
            SELECT quality_level, pi_q 
            FROM quality_factors 
            ORDER BY pi_q
        ''').fetchall()
        conn.close()
        
        return jsonify([dict(row) for row in qualities])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/environments')
def get_environments():
    """Get all environment types"""
    try:
        conn = get_db_connection()
        environments = conn.execute('''
            SELECT environment, pi_e 
            FROM environment_factors 
            ORDER BY environment
        ''').fetchall()
        conn.close()
        
        return jsonify([dict(row) for row in environments])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/resistor-styles')
def get_resistor_styles():
    """Get all resistor styles"""
    try:
        conn = get_db_connection()
        styles = conn.execute('''
            SELECT style, spec_number, description, lambda_b, 
                   pi_t_column, pi_s_column
            FROM resistor_styles 
            ORDER BY style
        ''').fetchall()
        conn.close()
        
        return jsonify([dict(row) for row in styles])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/resistor-quality-levels')
def get_resistor_quality_levels():
    """Get resistor quality levels"""
    try:
        conn = get_db_connection()
        qualities = conn.execute('''
            SELECT quality_level, pi_q 
            FROM resistor_quality_factors 
            ORDER BY pi_q
        ''').fetchall()
        conn.close()
        
        return jsonify([dict(row) for row in qualities])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/resistor-environments')
def get_resistor_environments():
    """Get resistor environment types"""
    try:
        conn = get_db_connection()
        environments = conn.execute('''
            SELECT environment, pi_e 
            FROM resistor_environment_factors 
            ORDER BY environment
        ''').fetchall()
        conn.close()
        
        return jsonify([dict(row) for row in environments])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/inductor-styles')
def get_inductor_styles():
    """Get all inductor styles"""
    try:
        conn = get_db_connection()
        styles = conn.execute('''
            SELECT inductor_type, lambda_b, pi_t_column
            FROM inductor_styles 
            ORDER BY inductor_type
        ''').fetchall()
        conn.close()
        
        return jsonify([dict(row) for row in styles])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/inductor-quality-levels')
def get_inductor_quality_levels():
    """Get inductor quality levels"""
    try:
        conn = get_db_connection()
        qualities = conn.execute('''
            SELECT quality_level, pi_q 
            FROM inductor_quality_factors 
            ORDER BY pi_q
        ''').fetchall()
        conn.close()
        
        return jsonify([dict(row) for row in qualities])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/inductor-environments')
def get_inductor_environments():
    """Get inductor environment types"""
    try:
        conn = get_db_connection()
        environments = conn.execute('''
            SELECT environment, pi_e 
            FROM inductor_environment_factors 
            ORDER BY environment
        ''').fetchall()
        conn.close()
        
        return jsonify([dict(row) for row in environments])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/calculate', methods=['POST'])
def calculate_reliability():
    """Calculate reliability for components"""
    try:
        data = request.get_json()
        components = data.get('components', [])
        
        if not components:
            return jsonify({'error': 'No components provided'}), 400
        
        results = []
        total_lambda_p = 0.0
        
        conn = get_db_connection()
        
        for component in components:
            # Determine component type
            component_type = component.get('component_type', 'capacitor')
            
            if component_type == 'resistor':
                result = calculate_resistor_reliability(conn, component)
            elif component_type == 'inductor':
                result = calculate_inductor_reliability(conn, component)
            else:
                result = calculate_component_reliability(conn, component)
            
            results.append(result)
            total_lambda_p += result['lambda_p']
        
        conn.close()
        
        return jsonify({
            'components': results,
            'total_lambda_p': round(total_lambda_p, 10),
            'calculation_timestamp': datetime.now().isoformat(),
            'component_count': len(results)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def calculate_resistor_temperature_factor(temperature, column):
    """Calculate resistor temperature factor using MIL-HDBK-217F equation"""
    if column == 1:
        ea = 0.2  # Column 1
    else:
        ea = 0.08  # Column 2
    
    temp_kelvin = temperature + 273
    factor = math.exp(-ea / Config.BOLTZMANN_CONSTANT * 
                     (1/temp_kelvin - 1/Config.REFERENCE_TEMP))
    return round(factor, 6)

def calculate_resistor_power_factor(power_dissipation):
    """Calculate resistor power factor using equation"""
    if power_dissipation <= 0:
        return 0.068
    
    # π_P = (Power Dissipation)^0.39
    factor = power_dissipation ** 0.39
    return round(factor, 7)

def calculate_resistor_stress_factor(power_stress, column):
    """Calculate resistor power stress factor"""
    if power_stress <= 0:
        return 0.66 if column == 2 else 0.79
    
    if column == 1:
        # Column 1: π_S = .71e^1.1(S)
        factor = 0.71 * math.exp(1.1 * power_stress)
    else:
        # Column 2: π_S = .54e^2.04(S)
        factor = 0.54 * math.exp(2.04 * power_stress)
    
    return round(factor, 6)

def calculate_inductor_temperature_factor(temperature):
    """Calculate inductor temperature factor using MIL-HDBK-217F equation"""
    # π_T = exp(-11 / (8.617 x 10^-5) * (1/(T+273) - 1/298))
    temp_kelvin = temperature + 273
    factor = math.exp(-11 / (8.617e-5) * (1/temp_kelvin - 1/298))
    return round(factor, 6)

def calculate_inductor_reliability(conn, component):
    """Calculate reliability for a single inductor component"""
    try:
        # Get component parameters
        project_name = component.get('project_name')
        inductor_type = component.get('inductor_type')
        temperature = float(component.get('temperature', Config.DEFAULT_TEMPERATURE))
        quality_level = component.get('quality_level', 'MIL-SPEC')
        environment = component.get('environment', Config.DEFAULT_ENVIRONMENT)
        
        # Enhanced component parameters
        component_name = component.get('name', 'Inductor_Component')
        description = component.get('description', '')
        manufacturer = component.get('manufacturer', '')
        part_number = component.get('part_number', '')
        
        # Get inductor style data
        style_data = conn.execute('''
            SELECT * FROM inductor_styles WHERE inductor_type = ?
        ''', (inductor_type,)).fetchone()
        
        if not style_data:
            raise ValueError(f"Inductor type '{inductor_type}' not found")
        
        lambda_b = style_data['lambda_b']
        
        # Calculate π_T (Temperature Factor)
        temp_data = conn.execute('''
            SELECT temperature, pi_t 
            FROM inductor_temperature_factors 
            ORDER BY temperature
        ''').fetchall()

        pi_t = get_exact_or_calculate_factor(temperature, temp_data, 'temperature', 'pi_t', 'inductor_temperature', None)
        
        # Get π_Q (Quality Factor)
        quality_data = conn.execute('''
            SELECT pi_q FROM inductor_quality_factors WHERE quality_level = ?
        ''', (quality_level,)).fetchone()
        
        if quality_data:
            pi_q = quality_data['pi_q']
        else:
            pi_q = 1.0
        
        # Get π_E (Environment Factor)
        env_data = conn.execute('''
            SELECT pi_e FROM inductor_environment_factors WHERE environment = ?
        ''', (environment,)).fetchone()
        
        if env_data:
            pi_e = env_data['pi_e']
        else:
            pi_e = 1.0
        
        # Calculate λ_P: λ_P = λ_b × π_T × π_Q × π_E
        lambda_p = lambda_b * pi_t * pi_q * pi_e
        
        return {
            'project_name': project_name,
            'name': component_name,
            'component_type': 'inductor',
            'style': inductor_type,
            'lambda_b': round(lambda_b, 8),
            'pi_t': round(pi_t, 6),
            'pi_q': round(pi_q, 6),
            'pi_e': round(pi_e, 6),
            'lambda_p': round(lambda_p, 10),
            'parameters': {
                'description': description,
                'manufacturer': manufacturer,
                'part_number': part_number,
                'temperature': temperature,
                'quality_level': quality_level,
                'environment': environment
            }
        }
        
    except Exception as e:
        raise Exception(f"Error calculating inductor {component.get('name', 'Unknown')}: {str(e)}")

def calculate_resistor_reliability(conn, component):
    """Calculate reliability for a single resistor component"""
    try:
        # Get component parameters
        project_name = component.get('project_name')
        style = component.get('style')
        temperature = float(component.get('temperature', Config.DEFAULT_TEMPERATURE))
        watts = float(component.get('watts', 0.125))  # Power dissipation in Watts
        power_stress = float(component.get('power_stress', 0.5))  # S value
        quality_level = component.get('quality_level', Config.DEFAULT_QUALITY)
        environment = component.get('environment', Config.DEFAULT_ENVIRONMENT)
        
        # Enhanced component parameters
        component_name = component.get('name', f'{style}_Component')
        description = component.get('description', '')
        manufacturer = component.get('manufacturer', '')
        part_number = component.get('part_number', '')
        
        # Get resistor style data
        style_data = conn.execute('''
            SELECT * FROM resistor_styles WHERE style = ?
        ''', (style,)).fetchone()
        
        if not style_data:
            raise ValueError(f"Resistor style '{style}' not found")
        
        lambda_b = style_data['lambda_b']
        pi_t_column = style_data['pi_t_column']
        pi_s_column = style_data['pi_s_column']
        
        # Calculate π_T (Temperature Factor)
        temp_data = conn.execute('''
            SELECT temperature, column_1, column_2 
            FROM resistor_temperature_factors 
            ORDER BY temperature
        ''').fetchall()

        if pi_t_column == 1:
            pi_t = get_exact_or_calculate_factor(temperature, temp_data, 'temperature', 'column_1', 'resistor_temperature', pi_t_column)
        else:
            pi_t = get_exact_or_calculate_factor(temperature, temp_data, 'temperature', 'column_2', 'resistor_temperature', pi_t_column)
        
        # Calculate π_P (Power Factor) using Watts
        power_data = conn.execute('''
            SELECT power_dissipation, pi_p 
            FROM resistor_power_factors 
            ORDER BY power_dissipation
        ''').fetchall()
        
        pi_p = get_exact_or_calculate_factor(watts, power_data, 'power_dissipation', 'pi_p', 'resistor_power', None)
        
        # Calculate π_S (Power Stress Factor) using S
        stress_data = conn.execute('''
            SELECT power_stress, column_1, column_2
            FROM resistor_stress_factors 
            ORDER BY power_stress
        ''').fetchall()

        column_name = f'column_{pi_s_column}'
        pi_s = get_exact_or_calculate_factor(power_stress, stress_data, 'power_stress', column_name, 'resistor_stress', pi_s_column)
        
        # Get π_Q (Quality Factor)
        quality_data = conn.execute('''
            SELECT pi_q FROM resistor_quality_factors WHERE quality_level = ?
        ''', (quality_level,)).fetchone()
        
        if quality_data:
            pi_q = quality_data['pi_q']
        else:
            pi_q = 3.0
        
        # Get π_E (Environment Factor)
        env_data = conn.execute('''
            SELECT pi_e FROM resistor_environment_factors WHERE environment = ?
        ''', (environment,)).fetchone()
        
        if env_data:
            pi_e = env_data['pi_e']
        else:
            pi_e = 1.0
        
        # Calculate λ_P: λ_P = λ_b × π_T × π_P × π_S × π_Q × π_E
        lambda_p = lambda_b * pi_t * pi_p * pi_s * pi_q * pi_e
        
        return {
            'project_name': project_name,
            'name': component_name,
            'component_type': 'resistor',
            'style': style,
            'lambda_b': round(lambda_b, 8),
            'pi_t': round(pi_t, 6),
            'pi_p': round(pi_p, 7),
            'pi_s': round(pi_s, 6),
            'pi_q': round(pi_q, 6),
            'pi_e': round(pi_e, 6),
            'lambda_p': round(lambda_p, 10),
            'parameters': {
                'description': description,
                'manufacturer': manufacturer,
                'part_number': part_number,
                'temperature': temperature,
                'watts': watts,
                'power_stress': power_stress,
                'quality_level': quality_level,
                'environment': environment
            }
        }
        
    except Exception as e:
        raise Exception(f"Error calculating resistor {component.get('name', 'Unknown')}: {str(e)}")

def calculate_component_reliability(conn, component):
    """Calculate reliability for a single component with enhanced parameters"""
    try:
        # Get component parameters
        project_name = component.get('project_name')
        style = component.get('style')
        temperature = float(component.get('temperature', Config.DEFAULT_TEMPERATURE))
        capacitance = float(component.get('capacitance', 1.0))
        voltage_stress = float(component.get('voltage_stress', 0.5))
        quality_level = component.get('quality_level', Config.DEFAULT_QUALITY)
        environment = component.get('environment', Config.DEFAULT_ENVIRONMENT)
        series_resistance = float(component.get('series_resistance', 1))
        
        # Enhanced component parameters
        component_name = component.get('name', f'{style}_Component')
        if not component_name or component_name == f'{style}_Component':
            component_type = 'Capacitor'
            component_name = f'{component_type}_{component.get("id", "Unknown")}'
        description = component.get('description', '')
        manufacturer = component.get('manufacturer', '')
        part_number = component.get('part_number', '')
        
        # Get capacitor style data
        style_data = conn.execute('''
            SELECT * FROM capacitor_styles WHERE style = ?
        ''', (style,)).fetchone()
        
        if not style_data:
            raise ValueError(f"Capacitor style '{style}' not found")
        
        lambda_b = style_data['lambda_b']
        pi_t_column = style_data['pi_t_column']
        pi_c_column = style_data['pi_c_column']
        pi_v_column = style_data['pi_v_column']
        default_pi_sr = style_data['pi_sr']
        
        # Calculate π_T (Temperature Factor)
        temp_data = conn.execute('''
            SELECT temperature, column_1, column_2 
            FROM temperature_factors 
            ORDER BY temperature
        ''').fetchall()

        if pi_t_column == 1:
            pi_t = get_exact_or_calculate_factor(temperature, temp_data, 'temperature', 'column_1', 'temperature', pi_t_column)
        else:
            pi_t = get_exact_or_calculate_factor(temperature, temp_data, 'temperature', 'column_2', 'temperature', pi_t_column)
        
        # Calculate π_C (Capacitance Factor)
        cap_data = conn.execute('''
            SELECT capacitance, column_1, column_2 
            FROM capacitance_factors 
            ORDER BY capacitance
        ''').fetchall()

        if pi_c_column == 1:
            pi_c = get_exact_or_calculate_factor(capacitance, cap_data, 'capacitance', 'column_1', 'capacitance', pi_c_column)
        else:
            pi_c = get_exact_or_calculate_factor(capacitance, cap_data, 'capacitance', 'column_2', 'capacitance', pi_c_column)
        
        # Calculate π_V (Voltage Stress Factor)
        voltage_data = conn.execute('''
            SELECT voltage_stress, column_1, column_2, column_3, column_4, column_5
            FROM voltage_stress_factors 
            ORDER BY voltage_stress
        ''').fetchall()

        column_name = f'column_{pi_v_column}'
        pi_v = get_exact_or_calculate_factor(voltage_stress, voltage_data, 'voltage_stress', column_name, 'voltage_stress', pi_v_column)
        
        # Get π_Q (Quality Factor)
        quality_data = conn.execute('''
            SELECT pi_q FROM quality_factors WHERE quality_level = ?
        ''', (quality_level,)).fetchone()
        
        if quality_data:
            pi_q = quality_data['pi_q']
        else:
            pi_q = 3.0  # Default for non-established reliability
        
        # Get π_E (Environment Factor)
        env_data = conn.execute('''
            SELECT pi_e FROM environment_factors WHERE environment = ?
        ''', (environment,)).fetchone()
        
        if env_data:
            pi_e = env_data['pi_e']
        else:
            pi_e = 1.0  # Default ground benign
        
        # Calculate π_SR (Series Resistance Factor) for tantalum capacitors
        pi_sr = default_pi_sr
        
        # Check if this is a tantalum capacitor that needs series resistance calculation
        tantalum_styles = ['CSR', 'CWR', 'CL', 'CLR', 'CRL']
        if any(ts in style.upper() for ts in tantalum_styles):
            # Determine resistance range based on series_resistance value
            if series_resistance > 0.8:
                pi_sr = 0.66
            elif series_resistance > 0.6:
                pi_sr = 1.0
            elif series_resistance > 0.4:
                pi_sr = 1.3
            elif series_resistance > 0.2:
                pi_sr = 2.0
            elif series_resistance > 0.1:
                pi_sr = 2.7
            else:
                pi_sr = 3.3
        
        # Calculate λ_P (Predicted Failure Rate)
        # λ_P = λ_b × π_T × π_C × π_V × π_Q × π_E × π_SR
        lambda_p = lambda_b * pi_t * pi_c * pi_v * pi_q * pi_e * pi_sr
        
        return {
            'project_name':project_name,
            'name': component_name,
            'style': style,
            'lambda_b': round(lambda_b, 8),
            'pi_t': round(pi_t, 6),
            'pi_c': round(pi_c, 6),
            'pi_v': round(pi_v, 6),
            'pi_q': round(pi_q, 6),
            'pi_e': round(pi_e, 6),
            'pi_sr': round(pi_sr, 6),
            'lambda_p': round(lambda_p, 10),
            'parameters': {
                'description': description,
                'manufacturer': manufacturer,
                'part_number': part_number,
                'temperature': temperature,
                'capacitance': capacitance,
                'voltage_stress': voltage_stress,
                'quality_level': quality_level,
                'environment': environment,
                'series_resistance': series_resistance
            }
        }
        
    except Exception as e:
        raise Exception(f"Error calculating component {component.get('name', 'Unknown')}: {str(e)}")

@app.route('/api/export/<format>')
def export_data(format):
    """Export calculation results - simplified for project-based approach"""
    try:
        if format not in ['json', 'csv']:
            return jsonify({'error': 'Invalid format. Use /api/project/export for project exports.'}), 400
        
        conn = get_db_connection()
        data = conn.execute('''
            SELECT * FROM calculations_history 
            ORDER BY created_at DESC
            LIMIT 50
        ''').fetchall()
        conn.close()
        
        if format == 'json':
            return jsonify([dict(row) for row in data])
            
    except Exception as e:
        print(f"Export Error: {str(e)}")
        return jsonify({'error': f'Export failed: {str(e)}'}), 500

def create_excel_export(project_data):
    """Create single-sheet Excel export with all information"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reliability Report"
    
    # Define WIB timezone (UTC+7)
    wib_tz = timezone(timedelta(hours=7))

    # Define styles
    title_font = Font(color="2563EB", bold=True, size=16)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    subheader_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def format_datetime_wib(iso_string):
        if not iso_string:
            return ''
        try:
            dt = datetime.fromisoformat(iso_string.replace('Z', '+00:00'))
            dt_wib = dt.astimezone(wib_tz)
            return dt_wib.strftime('%Y-%m-%d %H:%M:%S WIB')
        except:
            return iso_string

    current_row = 1
    
    # === SECTION 1: Title ===
    ws.cell(row=current_row, column=1, value="MIL-HDBK-217F Reliability Prediction Report")
    ws.cell(row=current_row, column=1).font = title_font
    ws.merge_cells(f'A{current_row}:K{current_row}')
    current_row += 2
    
    # === SECTION 2: Project Information ===
    ws.cell(row=current_row, column=1, value="PROJECT INFORMATION")
    ws.cell(row=current_row, column=1).font = subheader_font
    ws.cell(row=current_row, column=1).fill = subheader_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    project_info = [
        ("Project Name:", project_data.get('name', 'Untitled')),
        ("Description:", project_data.get('description', '')),
        ("Created Date:", format_datetime_wib(project_data.get('createdAt', ''))),
        ("Modified Date:", format_datetime_wib(project_data.get('modifiedAt', ''))),
        ("Version:", project_data.get('version', '1.1.0')),
        ("Temperature (°C):", project_data.get('globalParameters', {}).get('temperature', '')),
        ("Environment:", project_data.get('globalParameters', {}).get('environment', ''))
    ]
    
    for label, value in project_info:
        ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=value)
        current_row += 1
    
    current_row += 1
    
    # === SECTION 3: Components Configuration ===
    ws.cell(row=current_row, column=1, value="COMPONENTS CONFIGURATION")
    ws.cell(row=current_row, column=1).font = subheader_font
    ws.cell(row=current_row, column=1).fill = subheader_fill
    ws.merge_cells(f'A{current_row}:L{current_row}')
    current_row += 1

    components = project_data.get('components', [])

    # Detect component types in configuration
    comp_has_capacitors = any(c.get('component_type', 'capacitor') == 'capacitor' for c in components)
    comp_has_resistors = any(c.get('component_type') == 'resistor' for c in components)
    comp_has_inductors = any(c.get('component_type') == 'inductor' for c in components)

    # Build dynamic headers for components
    comp_headers = ['Component', 'Type', 'Description', 'Manufacturer', 'Part Number', 'Style/Type']

    if comp_has_capacitors:
        comp_headers.extend(['Cap (μF)', 'V.Stress', 'Series R (Ω)'])

    if comp_has_resistors:
        comp_headers.extend(['Watts', 'P.Stress'])

    comp_headers.append('Quality')

    # Write headers
    for col, header in enumerate(comp_headers, start=1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

    current_row += 1

    # Write component data rows
    for comp in components:
        comp_type = comp.get('component_type', 'capacitor')
        type_icon = '🔋' if comp_type == 'capacitor' else ('⚡' if comp_type == 'resistor' else '🔗')
        
        col_idx = 1
        
        # Component name
        ws.cell(row=current_row, column=col_idx, value=comp.get('name', ''))
        col_idx += 1
        
        # Type
        ws.cell(row=current_row, column=col_idx, value=f"{type_icon} {comp_type.capitalize()}")
        col_idx += 1
        
        # Description
        ws.cell(row=current_row, column=col_idx, value=comp.get('description', ''))
        col_idx += 1
        
        # Manufacturer
        ws.cell(row=current_row, column=col_idx, value=comp.get('manufacturer', ''))
        col_idx += 1
        
        # Part Number
        ws.cell(row=current_row, column=col_idx, value=comp.get('part_number', ''))
        col_idx += 1
        
        # Style/Type column
        if comp_type == 'inductor':
            ws.cell(row=current_row, column=col_idx, value=comp.get('inductor_type', ''))
        else:
            ws.cell(row=current_row, column=col_idx, value=comp.get('style', ''))
        col_idx += 1
        
        # Capacitor-specific parameters
        if comp_has_capacitors:
            if comp_type == 'capacitor':
                ws.cell(row=current_row, column=col_idx, value=comp.get('capacitance', ''))
                ws.cell(row=current_row, column=col_idx + 1, value=comp.get('voltage_stress', ''))
                ws.cell(row=current_row, column=col_idx + 2, value=comp.get('series_resistance', ''))
            else:
                ws.cell(row=current_row, column=col_idx, value='-')
                ws.cell(row=current_row, column=col_idx + 1, value='-')
                ws.cell(row=current_row, column=col_idx + 2, value='-')
            col_idx += 3
        
        # Resistor-specific parameters
        if comp_has_resistors:
            if comp_type == 'resistor':
                ws.cell(row=current_row, column=col_idx, value=comp.get('watts', ''))
                ws.cell(row=current_row, column=col_idx + 1, value=comp.get('power_stress', ''))
            else:
                ws.cell(row=current_row, column=col_idx, value='-')
                ws.cell(row=current_row, column=col_idx + 1, value='-')
            col_idx += 2
        
        # Quality (common for all)
        ws.cell(row=current_row, column=col_idx, value=comp.get('quality_level', ''))
        
        # Apply borders and alignment
        for col in range(1, col_idx + 1):
            ws.cell(row=current_row, column=col).border = border_thin
            ws.cell(row=current_row, column=col).alignment = Alignment(vertical='center')
        
        current_row += 1

    current_row += 1
    
    # === SECTION 4: Calculation Results ===
    if project_data.get('results'):
        ws.cell(row=current_row, column=1, value="CALCULATION RESULTS")
        ws.cell(row=current_row, column=1).font = subheader_font
        ws.cell(row=current_row, column=1).fill = subheader_fill
        ws.merge_cells(f'A{current_row}:N{current_row}')
        current_row += 1
        
        results = project_data['results']['components']
        
        # Detect component types
        has_capacitors = any(r.get('component_type') == 'capacitor' or not r.get('component_type') for r in results)
        has_resistors = any(r.get('component_type') == 'resistor' for r in results)
        has_inductors = any(r.get('component_type') == 'inductor' for r in results)
        
        # Build dynamic headers
        result_headers = ['Component', 'Type', 'Style', 'λb', 'πT']
        
        if has_capacitors:
            result_headers.extend(['πC', 'πV'])
        
        if has_resistors:
            result_headers.extend(['πP', 'πS'])
        
        result_headers.extend(['πQ', 'πE'])
        
        if has_capacitors:
            result_headers.append('πSR')
        
        result_headers.append('λP')
        
        # Write headers
        for col, header in enumerate(result_headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
        
        current_row += 1
        
        # Write data rows
        for result in results:
            component_type = result.get('component_type', 'capacitor')
            type_icon = '🔋' if component_type == 'capacitor' else ('⚡' if component_type == 'resistor' else '🔗')
            
            col_idx = 1
            
            # Component name
            ws.cell(row=current_row, column=col_idx, value=result.get('name', '-'))
            col_idx += 1
            
            # Type
            ws.cell(row=current_row, column=col_idx, value=f"{type_icon} {component_type.capitalize()}")
            col_idx += 1
            
            # Style
            if component_type == 'inductor':
                ws.cell(row=current_row, column=col_idx, value=result.get('inductor_type', result.get('style', '-')))
            else:
                ws.cell(row=current_row, column=col_idx, value=result.get('style', '-'))
            col_idx += 1
            
            # λb
            ws.cell(row=current_row, column=col_idx, value=result.get('lambda_b', '-'))
            col_idx += 1
            
            # πT
            ws.cell(row=current_row, column=col_idx, value=result.get('pi_t', '-'))
            col_idx += 1
            
            # πC and πV (capacitor only)
            if has_capacitors:
                if component_type == 'capacitor':
                    ws.cell(row=current_row, column=col_idx, value=result.get('pi_c', '-'))
                    ws.cell(row=current_row, column=col_idx + 1, value=result.get('pi_v', '-'))
                else:
                    ws.cell(row=current_row, column=col_idx, value='-')
                    ws.cell(row=current_row, column=col_idx + 1, value='-')
                col_idx += 2
            
            # πP and πS (resistor only)
            if has_resistors:
                if component_type == 'resistor':
                    ws.cell(row=current_row, column=col_idx, value=result.get('pi_p', '-'))
                    ws.cell(row=current_row, column=col_idx + 1, value=result.get('pi_s', '-'))
                else:
                    ws.cell(row=current_row, column=col_idx, value='-')
                    ws.cell(row=current_row, column=col_idx + 1, value='-')
                col_idx += 2
            
            # πQ
            ws.cell(row=current_row, column=col_idx, value=result.get('pi_q', '-'))
            col_idx += 1
            
            # πE
            ws.cell(row=current_row, column=col_idx, value=result.get('pi_e', '-'))
            col_idx += 1
            
            # πSR (capacitor only)
            if has_capacitors:
                if component_type == 'capacitor':
                    ws.cell(row=current_row, column=col_idx, value=result.get('pi_sr', '1.0'))
                else:
                    ws.cell(row=current_row, column=col_idx, value='-')
                col_idx += 1
            
            # λP
            lambda_p_value = result.get('lambda_p', '-')
            ws.cell(row=current_row, column=col_idx, value=str(lambda_p_value))
            
            # Apply borders
            for col in range(1, col_idx + 1):
                ws.cell(row=current_row, column=col).border = border_thin
                ws.cell(row=current_row, column=col).alignment = Alignment(vertical='center')
            
            current_row += 1
        
        # Total row
        total_col_span = len(result_headers) - 1
        ws.cell(row=current_row, column=1, value="TOTAL SYSTEM")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = subheader_fill
        ws.merge_cells(f'A{current_row}:{get_column_letter(total_col_span)}{current_row}')
        
        total_lambda_value = project_data['results'].get('total_lambda_p', '')
        ws.cell(row=current_row, column=len(result_headers), value=str(total_lambda_value))
        ws.cell(row=current_row, column=len(result_headers)).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=len(result_headers)).fill = subheader_fill
        
        for col in range(1, len(result_headers) + 1):
            ws.cell(row=current_row, column=col).border = border_thin
        
        current_row += 1

    # Auto-size columns
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    
    return wb

def parse_excel_import(file_stream):
    """Parse single-sheet Excel file and extract project data"""
    # Helper function to safely get cell value
    def safe_get_cell(row, col_key, default=''):
        try:
            col = col_indices.get(col_key)
            if col:
                val = ws.cell(row=row, column=col).value
                if val is None or str(val).strip() == '-':
                    return default
                return str(val).strip()
            return default
        except Exception as e:
            print(f"Error reading cell at row {row}, col {col_key}: {e}")
            return default
        
    wb = openpyxl.load_workbook(file_stream)
    ws = wb.active  # Ambil sheet pertama/aktif
    
    project_data = {
        'components': [],
        'globalParameters': {}
    }
    
    # Helper function untuk mencari baris berdasarkan keyword
    def find_row_with_keyword(keyword):
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and keyword.upper() in str(cell_value).upper():
                return row
        return None
    
    # === SECTION 1: Read Project Information ===
    project_info_row = find_row_with_keyword("PROJECT INFORMATION")
    
    if project_info_row:
        # Read project details starting from next row
        info_start = project_info_row + 1
        
        for row in range(info_start, info_start + 10):  # Max 10 rows untuk info
            label = ws.cell(row=row, column=1).value
            value = ws.cell(row=row, column=2).value
            
            if not label:
                break
                
            label_str = str(label).lower()
            
            if 'project name' in label_str:
                project_data['name'] = value or 'Imported Project'
            elif 'description' in label_str:
                project_data['description'] = value or ''
            elif 'version' in label_str:
                project_data['version'] = value or '1.1.0'
            elif 'temperature' in label_str:
                try:
                    project_data['globalParameters']['temperature'] = float(value) if value else 25
                except:
                    project_data['globalParameters']['temperature'] = 25
            elif 'environment' in label_str:
                project_data['globalParameters']['environment'] = value or 'GB'
    
    # === SECTION 2: Read Components Configuration ===
    comp_section_row = find_row_with_keyword("COMPONENTS CONFIGURATION")
    
    if comp_section_row:
        # Header row is next row after section title
        header_row = comp_section_row + 1
        
        # Find column indices from headers
        col_indices = {}
        for col in range(1, 20):
            header = ws.cell(row=header_row, column=col).value
            if not header:
                continue
            
            header_str = str(header).strip().lower()
            
            # Match exact order from export: Component, Type, Description, Manufacturer, Part Number, Style/Type, Cap, V.Stress, Watts, P.Stress, Quality, Series R
            if col_indices.get('name') is None and 'component' in header_str and 'type' not in header_str:
                col_indices['name'] = col
            elif col_indices.get('component_type') is None and header_str == 'type':
                col_indices['component_type'] = col
            elif col_indices.get('description') is None and header_str == 'description':
                col_indices['description'] = col
            elif col_indices.get('manufacturer') is None and header_str == 'manufacturer':
                col_indices['manufacturer'] = col
            elif col_indices.get('part_number') is None and ('part' in header_str or 'number' in header_str):
                col_indices['part_number'] = col
            elif col_indices.get('style_or_type') is None and ('style' in header_str or 'type' in header_str) and col > 5:
                col_indices['style_or_type'] = col
            elif col_indices.get('capacitance') is None and ('cap' in header_str and 'μf' in header_str):
                col_indices['capacitance'] = col
            elif col_indices.get('voltage_stress') is None and 'v.stress' in header_str:
                col_indices['voltage_stress'] = col
            elif col_indices.get('watts') is None and header_str == 'watts':
                col_indices['watts'] = col
            elif col_indices.get('power_stress') is None and header_str == 'p.stress':
                col_indices['power_stress'] = col
            elif col_indices.get('quality_level') is None and header_str == 'quality':
                col_indices['quality_level'] = col
            elif col_indices.get('series_resistance') is None and ('series' in header_str and 'r' in header_str):
                col_indices['series_resistance'] = col

        # Debug: Print found indices
        print("Found column indices:", col_indices)
        
        # Read component data rows
        data_start = header_row + 1
        
        for row in range(data_start, ws.max_row + 1):
            name_cell = ws.cell(row=row, column=col_indices.get('name', 1)).value
            
            # Stop if empty row or reached next section
            if not name_cell:
                break
                
            # Skip if this is a section header
            if any(keyword in str(name_cell).upper() for keyword in ['CALCULATION', 'RESULT', 'TOTAL']):
                break
            
            # Determine component type from Type column or guess from data
            type_cell = ws.cell(row=row, column=col_indices.get('component_type', 2)).value
            component_type = 'capacitor'  # default

            if type_cell:
                type_str = str(type_cell).lower()
                if 'resistor' in type_str or '⚡' in type_str:
                    component_type = 'resistor'
                elif 'inductor' in type_str or '🔗' in type_str:
                    component_type = 'inductor'
                elif 'capacitor' in type_str or '🔋' in type_str:
                    component_type = 'capacitor'

            # If no type cell, try to guess from available data
            if not type_cell:
                style_val = ws.cell(row=row, column=col_indices.get('style_or_type', 6)).value
                if style_val:
                    style_str = str(style_val).upper()
                    if any(ind_type in style_str for ind_type in ['FIXED', 'VARIABLE', 'FILM']):
                        component_type = 'inductor'
                    elif style_str.startswith('R'):
                        component_type = 'resistor'

            component = {
                'name': name_cell,
                'component_type': component_type,
                'description': safe_get_cell(row, 'description', ''),
                'manufacturer': safe_get_cell(row, 'manufacturer', ''),
                'part_number': safe_get_cell(row, 'part_number', ''),
                'temperature': project_data['globalParameters'].get('temperature', 25),
                'environment': project_data['globalParameters'].get('environment', 'GB')
            }

            # Get style or inductor_type
            style_or_type_val = ws.cell(row=row, column=col_indices.get('style_or_type', 6)).value
            if component_type == 'inductor':
                component['inductor_type'] = str(style_or_type_val).strip() if style_or_type_val and str(style_or_type_val).strip() != '-' else ''
            else:
                component['style'] = str(style_or_type_val).strip() if style_or_type_val and str(style_or_type_val).strip() != '-' else ''

            # Validation: Style/inductor_type must not be empty
            if component_type == 'capacitor' and not component.get('style'):
                print(f"Warning: Row {row} - Capacitor missing style, skipping")
                continue
            elif component_type == 'resistor' and not component.get('style'):
                print(f"Warning: Row {row} - Resistor missing style, skipping")
                continue
            elif component_type == 'inductor' and not component.get('inductor_type'):
                print(f"Warning: Row {row} - Inductor missing type, skipping")
                continue
            
            # Parse component-specific parameters
            if component_type == 'capacitor':
                try:
                    cap_value = ws.cell(row=row, column=col_indices.get('capacitance', 7)).value
                    component['capacitance'] = float(cap_value) if cap_value else 1.0
                except:
                    component['capacitance'] = 1.0
                
                try:
                    vs_value = ws.cell(row=row, column=col_indices.get('voltage_stress', 8)).value
                    component['voltage_stress'] = float(vs_value) if vs_value else 0.5
                except:
                    component['voltage_stress'] = 0.5
                
                try:
                    sr_value = ws.cell(row=row, column=col_indices.get('series_resistance', 12)).value
                    component['series_resistance'] = float(sr_value) if sr_value else 1.0
                except:
                    component['series_resistance'] = 1.0

            elif component_type == 'resistor':
                try:
                    watts_value = ws.cell(row=row, column=col_indices.get('watts', 9)).value
                    component['watts'] = float(watts_value) if watts_value else 0.125
                except:
                    component['watts'] = 0.125
                
                try:
                    ps_value = ws.cell(row=row, column=col_indices.get('power_stress', 10)).value
                    component['power_stress'] = float(ps_value) if ps_value else 0.5
                except:
                    component['power_stress'] = 0.5

            # Quality level (common for all types)
            quality_val = ws.cell(row=row, column=col_indices.get('quality_level', 11)).value
            if component_type == 'inductor':
                component['quality_level'] = quality_val or 'MIL-SPEC'
            else:
                component['quality_level'] = quality_val or 'M'
            
            project_data['components'].append(component)
    
    # Set defaults if not found
    if 'temperature' not in project_data['globalParameters']:
        project_data['globalParameters']['temperature'] = 25
    if 'environment' not in project_data['globalParameters']:
        project_data['globalParameters']['environment'] = 'GB'
    
    return project_data

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Endpoint not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """Export project to Excel format"""
    try:
        data = request.get_json()
        project_data = data.get('project')
        
        if not project_data:
            return jsonify({'error': 'No project data provided'}), 400
        
        wb = create_excel_export(project_data)
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename with WIB timezone
        wib_tz = timezone(timedelta(hours=7))
        now_wib = datetime.now(wib_tz)
        timestamp = now_wib.strftime('%Y%m%d_%H%M%S')
        
        project_name = project_data.get('name', 'project').replace(' ', '_')
        filename = f"{project_name}_{timestamp}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Excel Export Error: {str(e)}")
        return jsonify({'error': f'Excel export failed: {str(e)}'}), 500

@app.route('/api/import/excel', methods=['POST'])
def import_excel():
    """Import project from Excel format"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file'}), 400
        
        project_data = parse_excel_import(file)
        
        # Add metadata with WIB timezone
        wib_tz = timezone(timedelta(hours=7))
        now_wib = datetime.now(wib_tz)
        
        project_data['id'] = 'proj_' + now_wib.strftime('%Y%m%d%H%M%S')
        project_data['createdAt'] = now_wib.isoformat()
        project_data['modifiedAt'] = now_wib.isoformat()
        project_data['selectedComponentType'] = 'capacitor'
        
        return jsonify(project_data)
        
    except Exception as e:
        print(f"Excel Import Error: {str(e)}")
        return jsonify({'error': f'Excel import failed: {str(e)}'}), 500

if __name__ == '__main__':
    # Initialize database
    init_database()
    
    print(f"Starting Enhanced {Config.APP_NAME} v{Config.VERSION}")
    print(f"Database: {Config.get_database_path()}")
    print(f"Server: http://{Config.HOST}:{Config.PORT}")
    print(f"Features: Project-based workflow, Enhanced component parameters")
    
    app.run(
        host=Config.HOST,
        port=Config.PORT,
        debug=Config.DEBUG,
        threaded=True
    )